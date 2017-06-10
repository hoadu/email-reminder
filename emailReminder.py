# Send emails based on order status

import openpyxl
import smtplib

# Open the spreadsheet and get the latest status.

wb = openpyxl.load_workbook('customerStatus.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

# Check each member's pickup status.

unnotifiedCustomers = {}
for r in range(1, sheet.max_row + 1):
    status = sheet.cell(row=r, column=3).value
    sent = sheet.cell(row=r, column=4).value
    if status == 'yes' and sent != 'yes':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unnotifiedCustomers[name] = email
        sheet.cell(row=r, column=4).value = 'yes'
        wb.save('customerStatusTest.xlsx')

# Log in to email account

smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login(' youremailgoeshere@example.com ', 'yourpasswordgoeshere')

# Send out reminder emails

for name, email in unnotifiedCustomers.items():
    body = "Subject: -Enter subject here- \n-Enter the body here"
    print('Sending email to %s...' % email)
    mailStatus = smtpObj.sendmail(' youremailgoeshere@example.com ', email, body)

    if mailStatus != {}:
        print('There was a problem sending email to %s' % email)
    else:
        print('Email sent successfully.')
smtpObj.quit()
